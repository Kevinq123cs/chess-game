#!/usr/bin/env python3
"""Build a single-file financial-model visualizer from broker Excel models.

Usage:
    python3 build.py companies/weichai.py

Reads the company config (which Excel files, which labelled rows / cells to
pull), extracts the data, and injects it into template.html. The output is a
self-contained HTML file written to the repo root.

Rows are located by LABEL TEXT (normalized), not by row number, so the same
config style keeps working when a bank shifts rows around — and adapting to a
new company is mostly a matter of pointing at its files and fixing the few
labels that differ. SOTP sheets are free-form, so those use explicit cell
references.
"""
import importlib.util
import json
import math
import re
import sys
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent


def norm(s):
    return re.sub(r"\s+", " ", str(s)).strip().lower()


def clean(v, scale=1.0):
    if v is None or isinstance(v, str):
        return None
    try:
        f = float(v) * scale
    except (TypeError, ValueError):
        return None
    if math.isnan(f) or math.isinf(f):
        return None
    return round(f, 4) if abs(f) < 100 else round(f, 1)


class Model:
    """One bank's workbook with label-based row lookup."""

    def __init__(self, path):
        self.wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        self._cols = {}  # (sheet, col) -> [normalized labels]

    def _col(self, sheet, col):
        key = (sheet, col)
        if key not in self._cols:
            ws = self.wb[sheet]
            self._cols[key] = [
                norm(row[0]) if row[0] is not None else ""
                for row in ws.iter_rows(min_col=col, max_col=col, values_only=True)
            ]
        return self._cols[key]

    def find_row(self, sheet, col, label, occ=1, after=0):
        """Row number of the occ-th match of label in a column. Exact
        (whitespace/case-insensitive) match first, then unique startswith."""
        tgt = norm(label)
        vals = self._col(sheet, col)
        exact = [i + 1 for i, v in enumerate(vals) if v == tgt and i + 1 > after]
        if len(exact) >= occ:
            return exact[occ - 1]
        prefix = [i + 1 for i, v in enumerate(vals) if v.startswith(tgt) and v and i + 1 > after]
        if len(prefix) >= occ:
            return prefix[occ - 1]
        raise KeyError(f"label {label!r} (occ {occ}, after row {after}) not found in {sheet}!col{col}")

    def grab(self, sheet, row, col0, n, scale=1.0):
        ws = self.wb[sheet]
        vals = next(ws.iter_rows(min_row=row, max_row=row, min_col=col0,
                                 max_col=col0 + n - 1, values_only=True))
        return [clean(v, scale) for v in vals]

    def cell(self, sheet, ref):
        return self.wb[sheet][ref].value


def fmt_cells(model, sheet, template):
    """Replace {C4:,.0f}-style refs in a string with formatted cell values."""

    def repl(m):
        ref, spec = m.group(1), m.group(2) or ""
        v = model.cell(sheet, ref)
        if v is None:
            return "–"
        return format(v, spec) if spec else str(v)

    return re.sub(r"\{([A-Z]{1,3}\d+)(?::([^}]*))?\}", repl, template)


def pack(en, zh=None):
    """A display string: plain when monolingual, {en, zh} when translated."""
    return {"en": en, "zh": zh} if zh else en


def cagr(v0, v1, years):
    if not v0 or not v1 or v0 <= 0 or v1 <= 0 or years <= 0:
        return None
    return (v1 / v0) ** (1 / years) - 1


def auto_chips(series, compare):
    """Data-derived indicator chips for a deep-dive chart: evidence depth and
    cross-bank agreement. Authored (editorial) chips live on claims instead."""
    chips = []
    n_act = min(
        sum(1 for y, v in zip(s["years"], s["values"]) if v is not None and y < s["estFrom"])
        for s in series
    )
    chips.append({"type": "actuals", "n": n_act,
                  "tone": "bad" if n_act <= 1 else ("warn" if n_act <= 3 else "good")})
    if compare and len(compare) == 2:
        a, b = series[compare[0]], series[compare[1]]
        av = dict(zip(a["years"], a["values"]))
        bv = dict(zip(b["years"], b["values"]))
        est = [y for y in a["years"]
               if y >= a["estFrom"] and av.get(y) is not None and bv.get(y) is not None]
        if est:
            first, last = est[0], est[-1]
            g_first = av[first] / bv[first] - 1
            if abs(g_first) < 0.02:
                chips.append({"type": "aligned", "year": first, "tone": "good"})
            g_last = av[last] / bv[last] - 1
            tone = "good" if abs(g_last) < 0.10 else ("warn" if abs(g_last) < 0.30 else "bad")
            chips.append({"type": "gap", "year": last, "pct": round(g_last, 4), "tone": tone})
    return chips


def extract(cfg):
    models = {k: Model(s["file"]) for k, s in cfg["sources"].items()}

    sources = {}
    for key, s in cfg["sources"].items():
        y0, y1 = s["years"]
        sources[key] = {
            "name": pack(s["name"], s.get("nameZh")),
            "asOf": pack(s["asOf"], s.get("asOfZh")),
            "years": list(range(y0, y1 + 1)),
            "estFrom": s["estFrom"],
        }

    # ---- metrics (label-based row lookup) ----
    metrics = []
    for m in cfg["metrics"]:
        entry = {"id": m["id"], "label": pack(m["label"], m.get("labelZh")),
                 "unit": m["unit"], "category": m["category"], "series": {}}
        for key, spec in m["rows"].items():
            if spec is None:
                continue
            label, occ = (spec, 1) if isinstance(spec, str) else spec
            s = cfg["sources"][key]
            row = models[key].find_row(s["metricSheet"], s["labelCol"], label, occ)
            n = len(sources[key]["years"])
            entry["series"][key] = models[key].grab(s["metricSheet"], row, s["valueCol"], n)
        metrics.append(entry)

    # ---- segments ----
    # series entries are (displayEn, sheetLabel) or (displayEn, displayZh, sheetLabel)
    segments, seg_names = {}, {}
    for key, seg in cfg.get("segments", {}).items():
        model, out = models[key], {}
        n = len(sources[key]["years"])
        for gname, g in seg["groups"].items():
            after = 0
            for acol, alabel in g.get("anchors", []):
                after = model.find_row(seg["sheet"], acol, alabel, after=after)
            series = {}
            for item in g["series"]:
                display, zh, label = item if len(item) == 3 else (item[0], None, item[1])
                if zh:
                    seg_names[display] = zh
                row = model.find_row(seg["sheet"], seg["labelCol"], label, after=after)
                series[display] = model.grab(seg["sheet"], row, seg["valueCol"], n,
                                             scale=g.get("scale", 1.0))
            out[gname] = series
        out["notes"] = {k: pack(g["note"], g.get("noteZh"))
                        for k, g in seg["groups"].items() if g.get("note")}
        out["mixOf"] = seg.get("mixOf", "revenue")
        segments[key] = out

    # ---- SOTP (free-form sheets -> explicit cells) ----
    sotp = None
    if "sotp" in cfg:
        groups = [{"id": g["id"], "label": pack(g["label"], g.get("labelZh"))}
                  for g in cfg["sotp"]["groups"]]
        sotp = {"groups": groups, "sources": {}}
        for key, sp in cfg["sotp"]["sources"].items():
            model, sheet = models[key], sp["sheet"]
            parts = []
            for p in sp["parts"]:
                basis_en = fmt_cells(model, sheet, p["basis"])
                basis_zh = fmt_cells(model, sheet, p["basisZh"]) if p.get("basisZh") else None
                part = {
                    "name": pack(p["name"], p.get("nameZh")),
                    "group": p["group"],
                    "value": clean(model.cell(sheet, p["value"])),
                    "basis": pack(basis_en, basis_zh),
                }
                if p.get("old"):
                    part["old"] = clean(model.cell(sheet, p["old"]))
                parts.append(part)
            stats = [{"label": pack(st["label"], st.get("labelZh")),
                      "value": st["fmt"].format(v=model.cell(sheet, st["cell"]))}
                     for st in sp["stats"] if model.cell(sheet, st["cell"]) is not None]
            entry = {"parts": parts, "total": clean(model.cell(sheet, sp["total"])), "stats": stats}
            if sp.get("oldTotal"):
                entry["oldTotal"] = clean(model.cell(sheet, sp["oldTotal"]))
            sotp["sources"][key] = entry

    # ---- deep dives (driver-sheet detail; explicit rows, like SOTP cells) ----
    deep = []
    for dd in cfg.get("deepDives", []):
        charts = []
        for ch in dd.get("charts", []):
            series = []
            for sd in ch["series"]:
                model = models[sd["src"]]
                acc = None
                for r in (sd.get("sumRows") or [sd["row"]]):
                    vals = model.grab(sd["sheet"], r, sd["col0"], sd["n"], scale=sd.get("scale", 1.0))
                    if acc is None:
                        acc = vals
                    else:
                        acc = [None if (x is None and y is None) else (x or 0) + (y or 0)
                               for x, y in zip(acc, vals)]
                series.append({
                    "name": pack(sd["name"], sd.get("nameZh")),
                    "src": sd["src"], "slot": sd.get("slot", 1),
                    "years": list(range(sd["y0"], sd["y0"] + sd["n"])),
                    "values": acc,
                    "estFrom": cfg["sources"][sd["src"]]["estFrom"],
                })
            entry = {"title": pack(ch["title"], ch.get("titleZh")), "unit": ch["unit"],
                     "series": series, "chips": auto_chips(series, ch.get("compare"))}
            if ch.get("note"):
                entry["note"] = pack(ch["note"], ch.get("noteZh"))
            charts.append(entry)

        drivers = []
        for dr in dd.get("drivers", []):
            model, est_from = models[dr["src"]], cfg["sources"][dr["src"]]["estFrom"]
            rows_out = []
            for r in dr["rows"]:
                vals = model.grab(dr["sheet"], r["row"], dr["col0"], dr["n"], scale=r.get("scale", 1.0))
                pairs = [(y, v) for y, v in zip(range(dr["y0"], dr["y0"] + dr["n"]), vals)
                         if v is not None]
                act = [p for p in pairs if p[0] < est_from]
                est = [p for p in pairs if p[0] >= est_from]
                la, e1, eN = (act[-1] if act else None), (est[0] if est else None), (est[-1] if est else None)
                g = None
                if la and eN and not r.get("pct"):
                    g = cagr(la[1], eN[1], eN[0] - la[0])
                rows_out.append({"name": pack(r["name"], r.get("nameZh")),
                                 "unit": pack(r["unit"], r.get("unitZh")),
                                 "pct": r.get("pct", False),
                                 "lastA": la, "e1": e1, "eN": eN,
                                 "cagr": round(g, 4) if g is not None else None,
                                 "nActuals": len(act)})
            drivers.append({"src": dr["src"], "rows": rows_out})

        claims = [{"src": c["src"], "text": pack(c["text"], c.get("textZh")),
                   "chips": [{"k": pack(x["k"], x.get("kZh")), "v": pack(x["v"], x.get("vZh")),
                              "tone": x.get("tone", "neutral")} for x in c.get("chips", [])]}
                  for c in dd.get("claims", [])]
        deep.append({"id": dd["id"], "title": pack(dd["title"], dd.get("titleZh")),
                     "summaryKey": dd.get("summaryKey"),
                     "claims": claims, "charts": charts, "drivers": drivers})

    data = {
        "company": pack(cfg["company"], cfg.get("companyZh")),
        "tickers": cfg["tickers"],
        "currency": cfg["currency"],
        "sources": sources,
        "metrics": metrics,
        "segments": segments,
    }
    if seg_names:
        data["segNames"] = seg_names
    if cfg.get("categories"):
        data["categoryNames"] = {en: pack(en, zh) for en, zh in cfg["categories"].items()}
    if cfg.get("summaries"):
        data["summaries"] = cfg["summaries"]
    if sotp:
        data["sotp"] = sotp
    if deep:
        data["deepDives"] = deep
    return data


def main():
    if len(sys.argv) != 2:
        sys.exit(__doc__)
    cfg_path = (HERE / sys.argv[1]).resolve() if not Path(sys.argv[1]).is_absolute() else Path(sys.argv[1])
    spec = importlib.util.spec_from_file_location("company_cfg", cfg_path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    cfg = mod.CONFIG

    data = extract(cfg)
    payload = json.dumps(data, ensure_ascii=False)
    assert "</" not in payload, "JSON payload would break out of its <script> tag"

    template = (HERE / "template.html").read_text()
    assert "__DATA__" in template
    out_path = HERE.parent / cfg["output"]
    out_path.write_text(template.replace("__DATA__", payload))
    kb = out_path.stat().st_size / 1024
    print(f"wrote {out_path}  ({kb:.0f} KB, {len(data['metrics'])} metrics, "
          f"segments: {list(data['segments'])}, sotp: {list(data.get('sotp', {}).get('sources', {}))})")


if __name__ == "__main__":
    main()
